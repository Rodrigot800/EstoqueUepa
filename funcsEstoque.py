from openpyxl import load_workbook,Workbook
from datetime import datetime
from tkinter import filedialog
import os

def selecionar_e_preparar_planilha():
    caminho = filedialog.askopenfilename(
        title="Selecionar planilha de estoque",
        filetypes=[("Planilhas Excel", "*.xlsx")]
    )

    if not caminho:
        return None, None

    if os.path.exists(caminho):
        wb = load_workbook(caminho)
    else:
        wb = Workbook()

    # ---------------------------
    # ABA PRODUTOS
    # ---------------------------
    if "produtos" not in wb.sheetnames:
        ws_produtos = wb.create_sheet("produtos")
        ws_produtos.append([
            "ID",
            "Nome",
            "Unidade",
            "Qtd_Mínima",
            "Estoque"
        ])
    else:
        ws_produtos = wb["produtos"]

    # ---------------------------
    # ABA MOVIMENTOS
    # ---------------------------
    if "movimentos" not in wb.sheetnames:
        ws_mov = wb.create_sheet("movimentos")
        ws_mov.append([
            "ID",
            "ProdutoID",
            "ProdutoNome",
            "Tipo",
            "Quantidade",
            "Data"
        ])

    # Remove aba padrão se existir
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(caminho)
    wb.close()

    nome = os.path.basename(caminho)
    return caminho, nome

ARQUIVO = None

def set_arquivo(caminho):
    global ARQUIVO
    ARQUIVO = caminho

def carregar():
    if not ARQUIVO:
        raise Exception("Nenhuma planilha selecionada.")

    if not os.path.exists(ARQUIVO):
        raise Exception("Arquivo da planilha não encontrado.")

    return load_workbook(ARQUIVO)


# ----------------------------
# PRODUTOS
# ----------------------------
def listar_produtos():
    wb = carregar()
    ws = wb["produtos"]

    produtos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        produtos.append(row)

    wb.close()
    return produtos

def inserir_produto(nome, unidade, minimo):
    wb = carregar()
    ws = wb["produtos"]

    novo_id = ws.max_row
    ws.append([novo_id, nome, unidade, minimo, 0])

    wb.save(ARQUIVO)
    wb.close()

# ----------------------------
# MOVIMENTOS
# ----------------------------
def listar_movimentos():
    wb = carregar()
    ws = wb["movimentos"]

    movimentos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        movimentos.append(row)

    wb.close()
    return movimentos

def registrar_movimento(produto_id, produto_nome, tipo, quantidade):
    wb = carregar()
    ws = wb["movimentos"]

    novo_id = ws.max_row
    data = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.append([
        novo_id,
        produto_id,       # usado para estoque
        produto_nome,     # usado para histórico
        tipo,
        quantidade,
        data
    ])

    wb.save(ARQUIVO)
    wb.close()

    atualizar_estoque_produto(produto_id)


# ----------------------------
# SALDO
# ----------------------------

def atualizar_estoque_produto(produto_id):
    wb = carregar()
    ws = wb["produtos"]

    saldos = calcular_saldos()
    saldo = saldos.get(produto_id, 0)

    for row in ws.iter_rows(min_row=2):
        if row[0].value == produto_id:
            row[4].value = saldo  # coluna Estoque
            break

    wb.save(ARQUIVO)
    wb.close()

def inserir_produto(nome, unidade, minimo):
    wb = carregar()
    ws = wb["produtos"]

    novo_id = ws.max_row
    ws.append([novo_id, nome, unidade, minimo, 0])  # estoque começa 0

    wb.save(ARQUIVO)
    wb.close()

def calcular_saldos():
    saldos = {}

    for _, pid, _, tipo, qtd, _ in listar_movimentos():
        if pid is None:
            continue

        pid = int(pid)
        qtd = int(qtd)

        if pid not in saldos:
            saldos[pid] = 0

        if tipo == "ENTRADA":
            saldos[pid] += qtd
        elif tipo == "SAÍDA":
            saldos[pid] -= qtd

    return saldos


def aplicar_icone(janela):
    import os
    icone_path = os.path.join(
        os.path.dirname(__file__),
        "assets",
        "../assets/UepaEstoqueIcone.png"
    )
    janela.iconbitmap(icone_path)

def listar_movimentacoes():
    wb = load_workbook(ARQUIVO)
    ws = wb["movimentos"]  # ajuste se o nome da aba for outro

    dados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Exemplo de colunas:
        # data, produto_id, produto_nome, tipo, quantidade
        dados.append(row)

    wb.close()
    return dados